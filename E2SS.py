import io
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string


# ============================================================
# Profiles (mirrors JS PROFILES)
# ============================================================
PROFILES = [
    {
        "id": "cc_strict_row",
        "name": "Control Center Safe (Strict Row Formulas)",
        "opts": {
            "allowCrossSheetPlaceholders": False,
            "preferRowRefs": True,
            "allowRowNumberRefs": False,
            "strictUnsupportedFunctions": True,
            "rewriteNowToToday": True,
        },
    },
    {
        "id": "cc_strict_column",
        "name": "Control Center Safe (Strict Column Formulas)",
        "opts": {
            "allowCrossSheetPlaceholders": False,
            "preferRowRefs": False,
            "allowRowNumberRefs": False,
            "strictUnsupportedFunctions": True,
            "rewriteNowToToday": True,
        },
    },
    {
        "id": "cc_permissive",
        "name": "Control Center Safe (Permissive w/ Placeholders)",
        "opts": {
            "allowCrossSheetPlaceholders": True,
            "preferRowRefs": True,
            "allowRowNumberRefs": True,
            "strictUnsupportedFunctions": True,
            "rewriteNowToToday": True,
        },
    },
]

UNSUPPORTED_FUNCS = {"INDIRECT", "OFFSET"}
FLAG_FUNCS = {"SEQUENCE", "UNIQUE", "FILTER", "SORT", "SORTBY", "LET", "LAMBDA"}


# ============================================================
# Models
# ============================================================
@dataclass
class Notes:
    info: List[str] = field(default_factory=list)
    warn: List[str] = field(default_factory=list)
    error: List[str] = field(default_factory=list)


@dataclass
class Issues:
    errors: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)
    info: List[str] = field(default_factory=list)


@dataclass
class ConvertResult:
    converted: str
    notes: Notes
    issues: Issues
    has_lookup_template: bool


@dataclass
class Context:
    col_header_map: Dict[str, str]          # "A" -> "Header"
    target_row: int                         # row number of formula cell
    header_row: int
    profile_opts: dict
    output_mode: str                        # "row" or "column"
    strict_mode: bool
    # lookup mapping (set at runtime)
    lookup_value_header: str = ""
    match_header: str = ""
    return_header: str = ""
    optional_filter_header: str = ""


# ============================================================
# Helpers: header mapping
# ============================================================
def safe_header_name(v) -> Optional[str]:
    s = str(v).strip() if v is not None else ""
    return s if s else None


def build_col_header_map(ws, header_row_num: int) -> Tuple[Dict[str, str], List[str]]:
    """
    Mirrors buildColHeaderMap() in JS:
    - Uses worksheet max_column
    - Creates warnings for missing header names
    """
    warnings: List[str] = []
    out: Dict[str, str] = {}
    max_col = ws.max_column or 0
    for c in range(1, max_col + 1):
        col_letter = get_column_letter(c)
        header_val = ws.cell(row=header_row_num, column=c).value
        header = safe_header_name(header_val)
        if header:
            out[col_letter] = header
        else:
            warnings.append(f"Missing header for column {col_letter} (header row {header_row_num}).")
    return out, warnings


def parse_header_map_string(s: str) -> Dict[str, str]:
    """
    Mirrors parseHeaderMapString(): "A=Risk Score, B=Impact"
    """
    out: Dict[str, str] = {}
    if not s:
        return out
    for pair in s.split(","):
        if "=" not in pair:
            continue
        col, header = pair.split("=", 1)
        col = col.strip().upper()
        header = header.strip()
        if col and header:
            out[col] = header
    return out


def generic_header_for_col(col_letter: str) -> str:
    n = column_index_from_string(col_letter)
    return f"Col {n}"


# ============================================================
# Conversion: function rewrites (mirrors rewriteFunctions)
# ============================================================
def rewrite_functions(formula: str, opts: dict, notes: Notes) -> str:
    f = (formula or "").strip()
    if f.startswith("="):
        f = f[1:].strip()

    # NOW() -> TODAY()
    if opts.get("rewriteNowToToday", False):
        def _now_repl(_m):
            notes.warn.append("Rewrote NOW() to TODAY() (Smartsheet NOW differs from Excel NOW).")
            return "TODAY()"
        f = re.sub(r"\bNOW\s*\(\s*\)", _now_repl, f, flags=re.IGNORECASE)

    # COUNTIF(range, crit) -> COUNTIFS(range, crit)
    def _countif(m):
        rng, crit = m.group(1), m.group(2)
        notes.info.append("Rewrote COUNTIF(...) to COUNTIFS(...).")
        return f"COUNTIFS({rng}, {crit})"
    f = re.sub(r"\bCOUNTIF\s*\(\s*([^,]+)\s*,\s*([^)]+?)\s*\)", _countif, f, flags=re.IGNORECASE)

    # SUMIF(range, criteria, sum_range) -> SUMIFS(sum_range, range, criteria)
    def _sumif3(m):
        rng, crit, sumrng = m.group(1), m.group(2), m.group(3)
        notes.info.append("Rewrote SUMIF(range, criteria, sum_range) to SUMIFS(sum_range, range, criteria).")
        return f"SUMIFS({sumrng}, {rng}, {crit})"
    f = re.sub(r"\bSUMIF\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*,\s*([^)]+?)\s*\)", _sumif3, f, flags=re.IGNORECASE)

    # SUMIF(range, criteria) -> SUMIFS(range, range, criteria)
    def _sumif2(m):
        rng, crit = m.group(1), m.group(2)
        notes.warn.append("Rewrote SUMIF(range, criteria) to SUMIFS(range, range, criteria). Verify sum_range intent.")
        return f"SUMIFS({rng}, {rng}, {crit})"
    f = re.sub(r"\bSUMIF\s*\(\s*([^,]+)\s*,\s*([^)]+?)\s*\)", _sumif2, f, flags=re.IGNORECASE)

    # AVERAGEIF(range, criteria [, avg_range]) -> AVG(COLLECT(avg_range or range, range, criteria))
    def _avgif(m):
        rng, crit, avgrng = m.group(1), m.group(2), m.group(3)
        r = (avgrng.strip() if avgrng else "").strip()
        r = r if r else rng
        notes.warn.append("Rewrote AVERAGEIF(...) to AVG(COLLECT(...)) template. Validate range alignment.")
        return f"AVG(COLLECT({r}, {rng}, {crit}))"
    f = re.sub(r"\bAVERAGEIF\s*\(\s*([^,]+)\s*,\s*([^,]+)\s*(?:,\s*([^)]+?)\s*)?\)", _avgif, f, flags=re.IGNORECASE)

    # CONCATENATE(a,b,c) -> (a + b + c) with naive quoted-string aware split
    if re.search(r"\bCONCATENATE\s*\(", f, flags=re.IGNORECASE):
        def _concat(m):
            args = m.group(1)
            parts = re.findall(r'"(?:[^"]|"")*"|[^,]+', args) or []
            notes.warn.append("CONCATENATE normalized to '+' string concatenation for Smartsheet.")
            return "(" + " + ".join(p.strip() for p in parts) + ")"
        f = re.sub(r"\bCONCATENATE\s*\(([^()]*)\)", _concat, f, flags=re.IGNORECASE)

    # LOOKUP functions -> sentinel token to be finalized later
    def _lookup(m):
        fn = m.group(1).upper()
        args = m.group(2).replace('"', r'\"')
        notes.warn.append(f"{fn} detected. Converting to INDEX(COLLECT()) template. Use Lookup Mapping UI to finalize.")
        return f'__LOOKUP_TEMPLATE__("{fn}", "{args}")'
    f = re.sub(r"\b(XLOOKUP|VLOOKUP|HLOOKUP)\s*\(([\s\S]*?)\)", _lookup, f, flags=re.IGNORECASE)

    return "=" + f


# ============================================================
# Conversion: structured refs
# ============================================================
def replace_structured_refs(formula: str, notes: Notes) -> str:
    f = str(formula or "")

    # [@[Column]] -> [Column]@row
    def _this_row(m):
        col = m.group(1).strip()
        notes.warn.append(f"Converted Excel structured ref {m.group(0)} → [{col}]@row")
        return f"[{col}]@row"
    f = re.sub(r"\[\@\[(.+?)\]\]", _this_row, f)

    # [@Column] -> [Column]@row
    def _at_col(m):
        col = m.group(1).strip()
        notes.warn.append(f"Converted Excel structured ref {m.group(0)} → [{col}]@row")
        return f"[{col}]@row"
    f = re.sub(r"\[\@([^\]]+)\]", _at_col, f)

    return f


# ============================================================
# Conversion: A1 refs -> Smartsheet refs
# Mirrors replaceRefs + convertA1RangesToColumnRanges
# ============================================================
_A1_WHOLE_COL = re.compile(r"^(\$?[A-Z]{1,3})\s*:\s*(\$?[A-Z]{1,3})\b")
_A1_CELL = re.compile(r"^(\$?[A-Z]{1,3})(\$?\d+)\b")
_A1_RANGE = re.compile(r"^(\$?[A-Z]{1,3})(\$?\d+)\s*:\s*(\$?[A-Z]{1,3})(\$?\d+)\b")


def _col_header(ctx: Context, col_letter: str) -> Optional[str]:
    return ctx.col_header_map.get(col_letter)


def convert_a1_ranges_to_column_ranges(formula: str, ctx: Context, notes: Notes) -> str:
    s = str(formula or "")
    out = []
    in_str = False
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == '"':
            in_str = not in_str
            out.append(ch)
            i += 1
            continue
        if in_str:
            out.append(ch)
            i += 1
            continue

        sub = s[i:]
        m = _A1_RANGE.match(sub)
        if m:
            c1 = m.group(1).replace("$", "")
            c2 = m.group(3).replace("$", "")
            if c1 == c2:
                header = _col_header(ctx, c1)
                effective = header or generic_header_for_col(c1)
                if not header:
                    notes.warn.append(f"Used generic column mapping for range {m.group(0)} → [{effective}]:[{effective}].")
                out.append(f"[{effective}]:[{effective}]")
            else:
                out.append(m.group(0))
                notes.warn.append(f"Multi-column range {m.group(0)} not supported in Smartsheet.")
            i += len(m.group(0))
            continue

        out.append(ch)
        i += 1
    return "".join(out)


def replace_refs(formula: str, ctx: Context, notes: Notes) -> str:
    s = str(formula or "")
    out = []
    in_str = False
    i = 0
    while i < len(s):
        ch = s[i]
        if ch == '"':
            in_str = not in_str
            out.append(ch)
            i += 1
            continue
        if in_str:
            out.append(ch)
            i += 1
            continue

        sub = s[i:]

        # Whole-column A:A
        m = _A1_WHOLE_COL.match(sub)
        if m:
            left = m.group(1).replace("$", "")
            right = m.group(2).replace("$", "")
            if left == right:
                header = _col_header(ctx, left)
                effective = header or generic_header_for_col(left)
                if not header:
                    notes.warn.append(f"Used generic column mapping for range {m.group(0)} → [{effective}]:[{effective}].")
                out.append(f"[{effective}]:[{effective}]")
                i += len(m.group(0))
                continue

        # Cell A1
        m = _A1_CELL.match(sub)
        if m:
            col = m.group(1).replace("$", "")
            row = int(m.group(2).replace("$", ""))
            header = _col_header(ctx, col)
            effective = header or generic_header_for_col(col)

            if row == ctx.target_row:
                out.append(f"[{effective}]@row")
            elif ctx.profile_opts.get("allowRowNumberRefs", False):
                out.append(f"[{effective}]{row}")
                if not header:
                    notes.warn.append(f"Used generic column mapping for {col}{row} → [{effective}]{row}.")
            else:
                out.append(f"[{effective}]@row")
                notes.warn.append(f"Row number ref not allowed; coerced {col}{row} to @row using [{effective}].")

            i += len(m.group(0))
            continue

        out.append(ch)
        i += 1

    out_s = "".join(out)
    return convert_a1_ranges_to_column_ranges(out_s, ctx, notes)


# ============================================================
# Lookup template finalization
# ============================================================
_LOOKUP_SENTINEL = re.compile(r'__LOOKUP_TEMPLATE__\("([^"]+)",\s*"([^"]*)"\)')

def build_lookup_template(ctx: Context) -> Optional[str]:
    hv = ctx.lookup_value_header
    hm = ctx.match_header
    hr = ctx.return_header
    hf = ctx.optional_filter_header
    if not (hv and hm and hr):
        return None

    lookup_val = f"[{hv}]@row"
    match_ref = f"{{{hm}}}"
    return_ref = f"{{{hr}}}"

    if hf:
        filter_ref = f"{{{hf}}}"
        return f'INDEX(COLLECT({return_ref}, {match_ref}, {lookup_val}, {filter_ref}, "__FILTER_CRITERIA__"), 1)'
    return f"INDEX(COLLECT({return_ref}, {match_ref}, {lookup_val}), 1)"


def finalize_lookup_templates(formula: str, ctx: Context, notes: Notes) -> str:
    if "__LOOKUP_TEMPLATE__(" not in formula:
        return formula

    tpl = build_lookup_template(ctx)
    if not tpl:
        notes.warn.append("LOOKUP template present but no mapping selected. Emitting placeholders.")
        return _LOOKUP_SENTINEL.sub("INDEX(COLLECT(__MAP_RETURN__, __MAP_MATCH__, __MAP_LOOKUP__), 1)", formula)

    notes.info.append("Applied Lookup Mapping UI selections to LOOKUP templates.")
    return _LOOKUP_SENTINEL.sub(tpl, formula)


def apply_row_vs_column_preference(formula: str, mode: str, notes: Notes) -> str:
    if mode == "column":
        if re.search(r"@row\b", formula) and re.search(r"\b(SUM|COUNT|AVG|MIN|MAX|COUNTIFS|SUMIFS)\s*\(", formula, flags=re.IGNORECASE):
            notes.warn.append("Column-mode selected: formula uses @row inside an aggregation. Consider converting to Column:Column ranges.")
    return formula


# ============================================================
# Validator (mirrors validateSmartsheetFormula)
# ============================================================
def validate_smartsheet_formula(formula: str, profile_opts: dict, strict_mode: bool) -> Issues:
    issues = Issues()
    f = (formula or "").strip()

    if not f.startswith("="):
        issues.warnings.append("Formula does not start with '='. Smartsheet formula cells should start with '='.")

    for raw in re.findall(r"\b[A-Z_][A-Z0-9_]*\s*\(", f, flags=re.IGNORECASE):
        fn = raw.replace("(", "").strip().upper()
        if fn in UNSUPPORTED_FUNCS:
            issues.errors.append(f"Unsupported function: {fn} (Smartsheet Gov).")
        if fn in FLAG_FUNCS:
            issues.warnings.append(f"Excel function {fn} may not behave the same in Smartsheet; manual review recommended.")

    if re.search(r"\[[^\]]+\.xlsx\]", f, flags=re.IGNORECASE):
        issues.warnings.append("External workbook reference detected ([Book.xlsx]...). Smartsheet import will not preserve this.")

    if re.search(r"{[^}]+}", f):
        if not profile_opts.get("allowCrossSheetPlaceholders", False):
            issues.warnings.append("Cross-sheet reference placeholder detected ({Sheet Column}). Profile discourages placeholders; consider redesign.")
        else:
            issues.info.append("Cross-sheet placeholder present; ensure you create cross-sheet references in Smartsheet after import.")

    # Unbalanced parentheses check (string-safe)
    bal = 0
    in_str = False
    for ch in f:
        if ch == '"':
            in_str = not in_str
            continue
        if in_str:
            continue
        if ch == "(":
            bal += 1
        elif ch == ")":
            bal -= 1
            if bal < 0:
                break
    if bal != 0:
        issues.errors.append("Unbalanced parentheses detected.")

    if re.search(r"@cell\b", f, flags=re.IGNORECASE):
        issues.errors.append("@cell detected. Smartsheet Gov does not use @cell; use @row or column ranges.")

    if "__LOOKUP_TEMPLATE__(" in f:
        issues.warnings.append("Unresolved LOOKUP template remains. Use Lookup Mapping UI or complete manually.")

    if strict_mode and profile_opts.get("strictUnsupportedFunctions", False):
        if re.search(r"{\s*\d+\s*,", f):
            issues.warnings.append("Array constant detected (e.g., {1,2,3}). Smartsheet may not support this as expected.")

    return issues


# ============================================================
# Master conversion (mirrors convertExcelFormulaToSmartsheet)
# ============================================================
def convert_excel_formula_to_smartsheet(excel_formula: str, ctx: Context) -> ConvertResult:
    notes = Notes()
    f = (excel_formula or "").strip()

    if not f:
        return ConvertResult(
            converted="",
            notes=notes,
            issues=Issues(warnings=["Blank formula."]),
            has_lookup_template=False,
        )

    # Step 1: function rewrites (introduces lookup templates)
    f = rewrite_functions(f, ctx.profile_opts, notes)
    has_lookup_template = "__LOOKUP_TEMPLATE__(" in f

    # Step 1b: structured refs
    f = replace_structured_refs(f, notes)

    # Step 2: refs
    f = replace_refs(f, ctx, notes)

    # Step 3: finalize lookup templates (or emit placeholders)
    f = finalize_lookup_templates(f, ctx, notes)

    # Step 4: best-practice alignment warning
    f = apply_row_vs_column_preference(f, ctx.output_mode, notes)

    # Step 5: validate
    issues = validate_smartsheet_formula(f, ctx.profile_opts, ctx.strict_mode)

    return ConvertResult(
        converted=f,
        notes=notes,
        issues=issues,
        has_lookup_template=has_lookup_template,
    )


# ============================================================
# Batch: extract formulas from sheet
# ============================================================
def iter_formula_cells(ws):
    """
    Yields (address, row, col, formula_with_equals)
    openpyxl stores formulas in cell.value as '=...'
    """
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.startswith("=") and len(v.strip()) > 1:
                yield (cell.coordinate, cell.row, cell.column, v.strip())


# ============================================================
# Export helpers
# ============================================================
def build_smartsheet_ready_workbook(
    in_wb,
    sheet_name: str,
    results: List[dict],
    create_report: bool,
    strict_mode: bool,
) -> bytes:
    """
    Output workbook:
    - Target sheet: write converted formulas as TEXT (cell.data_type='s') (mirrors JS: cell.v=f, cell.t='s', delete cell.f)
    - Optional "Conversion Report" sheet
    """
    out_wb = Workbook()
    # remove default
    out_wb.remove(out_wb.active)

    in_ws = in_wb[sheet_name]
    out_ws = out_wb.create_sheet(title=sheet_name)

    # Copy values only (style parity not required)
    for r in range(1, in_ws.max_row + 1):
        for c in range(1, in_ws.max_column + 1):
            out_ws.cell(row=r, column=c, value=in_ws.cell(row=r, column=c).value)

    # Apply converted formulas as TEXT
    for r in results:
        addr = r["address"]
        cell = out_ws[addr]
        if strict_mode and r["errors"]:
            cell.value = r["original"]  # keep original formula as text
        else:
            f = (r["converted"] or "").strip()
            if f.startswith("'"):
                f = f[1:]
            cell.value = f  # text
        cell.data_type = "s"

    if create_report:
        rep = out_wb.create_sheet(title="Conversion Report")
        rep.append(["Cell","Row","Col","Original Excel Formula","Converted Smartsheet Formula","Status","Errors","Warnings/Notes"])
        for r in results:
            status = "BLOCKED (Strict)" if (strict_mode and r["errors"]) else ("CONVERTED (Warnings)" if r["warnings"] else "CONVERTED")
            rep.append([
                r["address"], r["row"], r["col"], r["original"], r["converted"], status,
                " | ".join(r["errors"]),
                " | ".join(r["warnings"]),
            ])

    bio = io.BytesIO()
    out_wb.save(bio)
    return bio.getvalue()


def build_paste_ready_csv(results: List[dict]) -> bytes:
    header = ["Cell", "Row", "Col", "Converted Smartsheet Formula", "Errors", "Warnings"]
    lines = [",".join(_csv_escape(x) for x in header)]
    for r in results:
        row = [
            r["address"],
            str(r["row"]),
            str(r["col"]),
            r["converted"] or "",
            " | ".join(r["errors"]),
            " | ".join(r["warnings"]),
        ]
        lines.append(",".join(_csv_escape(x) for x in row))
    return ("\n".join(lines)).encode("utf-8")


def _csv_escape(val: str) -> str:
    s = "" if val is None else str(val)
    if any(ch in s for ch in [",", '"', "\n", "\r"]):
        return '"' + s.replace('"', '""') + '"'
    return s


# ============================================================
# Streamlit UI
# ============================================================
st.set_page_config(page_title="Excel → Smartsheet Gov Formula Converter", layout="wide")
st.title("Excel → Smartsheet Gov Formula Converter (Python + Streamlit)")
st.caption("ingle-formula + single-sheet Excel intake, Gov-safe validation, lookup mapping, XLSX/CSV export.")

tab1, tab2 = st.tabs(["Single Formula Conversion", "Excel Intake (Single-Sheet)"])

def profile_by_id(pid: str) -> dict:
    for p in PROFILES:
        if p["id"] == pid:
            return p
    return PROFILES[0]


# ----------------------------
# Single formula
# ----------------------------
with tab1:
    colA, colB = st.columns([1, 1])
    with colA:
        excel_formula = st.text_area("Excel formula", placeholder='Example: =IF(COUNTIF(A:A,"High")>0, SUMIF(A:A,"High",B:B), "")', height=140)
        header_map_str = st.text_input("Column Header Mapping (optional)", placeholder="Example: A=Risk Score, B=Impact, C=Probability")
        st.caption("Used to convert A1 references into Smartsheet syntax.")

    with colB:
        context_row = st.number_input("Context Row Number (for @row mapping)", min_value=1, value=2, step=1)
        header_row = st.number_input("Header Row Number (for column name mapping)", min_value=1, value=1, step=1)
        prof_id = st.selectbox("Import Profile", [p["id"] for p in PROFILES], format_func=lambda x: profile_by_id(x)["name"])
        output_mode = st.selectbox("Output Mode", ["row", "column"], index=0, format_func=lambda x: "Prefer Row Formulas (@row)" if x=="row" else "Prefer Column Formulas (Column:Column)")
        strict = True

    if st.button("Convert", type="primary", disabled=not bool(excel_formula.strip())):
        col_header_map = parse_header_map_string(header_map_str)
        prof = profile_by_id(prof_id)
        ctx = Context(
            col_header_map=col_header_map,
            target_row=int(context_row),
            header_row=int(header_row),
            profile_opts={**prof["opts"], "preferRowRefs": output_mode == "row"},
            output_mode=output_mode,
            strict_mode=strict,
        )
        res = convert_excel_formula_to_smartsheet(excel_formula, ctx)

        st.subheader("Smartsheet formula")
        st.code(res.converted or "", language="text")

        used_generic = any("generic column mapping" in w.lower() for w in res.notes.warn)

        if res.issues.errors:
            st.error("BLOCK (Errors)")
        elif used_generic or res.issues.warnings or res.notes.warn:
            st.warning("WARN (Review recommended)")
        else:
            st.success("OK")

        msgs = []
        msgs += [f"ERROR: {e}" for e in res.issues.errors]
        msgs += [f"WARN: {w}" for w in res.issues.warnings]
        msgs += [f"NOTE: {w}" for w in res.notes.warn]
        msgs += [f"INFO: {i}" for i in res.issues.info]
        msgs += [f"INFO: {i}" for i in res.notes.info]
        st.text("\n".join(msgs) if msgs else "—")


# ----------------------------
# Batch intake
# ----------------------------
with tab2:
    st.write("Upload an .xlsx, select a sheet, analyze formulas, optionally map LOOKUP templates, then export XLSX/CSV.")

    up = st.file_uploader("Upload Excel (.xlsx)", type=["xlsx"])
    if "analysis" not in st.session_state:
        st.session_state.analysis = None

    if up:
        xlsx_bytes = up.getvalue()
        in_wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=False)

        left, right = st.columns([1, 1])
        with left:
            sheet_name = st.selectbox("Sheet", in_wb.sheetnames)
            header_row = st.number_input("Header Row Number", min_value=1, value=1, step=1, key="batch_header_row")
            prof_id = st.selectbox("Control Center–Safe Import Profile", [p["id"] for p in PROFILES], format_func=lambda x: profile_by_id(x)["name"], key="batch_profile")
            output_mode = st.selectbox("Row vs Column Best-Practice Alignment", ["row", "column"], index=0, key="batch_output_mode",
                                      format_func=lambda x: "Prefer Row Formulas (@row)" if x=="row" else "Prefer Column Formulas (Column:Column)")
        with right:
            strict_mode = st.checkbox("Strict Smartsheet Gov validation", value=True)
            create_report = st.checkbox('Create "Conversion Report" sheet', value=True)

        if st.button("Analyze", type="primary"):
            ws = in_wb[sheet_name]
            col_header_map, header_warnings = build_col_header_map(ws, int(header_row))

            prof = profile_by_id(prof_id)
            results = []
            for addr, r, c, formula in iter_formula_cells(ws):
                ctx = Context(
                    col_header_map=col_header_map,
                    target_row=r,
                    header_row=int(header_row),
                    profile_opts={**prof["opts"], "preferRowRefs": output_mode == "row"},
                    output_mode=output_mode,
                    strict_mode=bool(strict_mode),
                )
                res = convert_excel_formula_to_smartsheet(formula, ctx)
                results.append({
                    "address": addr,
                    "row": r,
                    "col": c,
                    "original": formula,
                    "converted": res.converted,
                    "originalRaw": re.sub(r"^=\s*", "", formula or ""),
                    "hasLookupTemplate": bool(res.has_lookup_template),
                    "errors": res.issues.errors,
                    "warnings": res.issues.warnings + res.notes.warn,
                    "info": res.issues.info + res.notes.info,
                })

            warn_count = sum(1 for r in results if r["warnings"])
            err_count = sum(1 for r in results if r["errors"])
            has_lookup_or_placeholders = any(r["hasLookupTemplate"] or re.search(r"__MAP_(RETURN|MATCH|LOOKUP)__", r["converted"] or "", flags=re.I) for r in results)

            st.session_state.analysis = {
                "workbook_name": up.name,
                "sheet_name": sheet_name,
                "header_row": int(header_row),
                "profile": profile_by_id(prof_id),
                "output_mode": output_mode,
                "strict_mode": bool(strict_mode),
                "create_report": bool(create_report),
                "col_header_map": col_header_map,
                "header_warnings": header_warnings,
                "results": results,
                "has_lookup_or_placeholders": has_lookup_or_placeholders,
                "warn_count": warn_count,
                "err_count": err_count,
            }

    analysis = st.session_state.analysis
    if analysis:
        st.divider()
        st.markdown("### Summary")
        c1, c2, c3, c4 = st.columns(4)
        c1.metric("Workbook", analysis["workbook_name"])
        c2.metric("Sheet", analysis["sheet_name"])
        c3.metric("Formulas found", len(analysis["results"]))
        c4.metric("Warnings / Errors", f'{analysis["warn_count"]} / {analysis["err_count"]}')

        if analysis["header_warnings"]:
            st.warning("Header issues (from selected header row):\n\n" + "\n".join(f"- {w}" for w in analysis["header_warnings"]))

        df = pd.DataFrame([
            {
                "Cell": r["address"],
                "Row": r["row"],
                "Col": r["col"],
                "Converted": r["converted"],
                "Errors": " | ".join(r["errors"]),
                "Warnings": " | ".join(r["warnings"]),
                "HasLookupTemplate": r["hasLookupTemplate"],
            }
            for r in analysis["results"]
        ])
        st.dataframe(df, use_container_width=True, height=360)

        # Lookup mapping UI
        st.markdown("### Lookup Mapping UI (for INDEX(COLLECT))")
        headers = list(dict.fromkeys([h for h in analysis["col_header_map"].values() if h]))  # dedupe

        lk1, lk2 = st.columns(2)
        with lk1:
            lookup_value_header = st.selectbox("Lookup Value Column (same row, usually @row)", [""] + headers, index=0)
            match_header = st.selectbox("Match Column (range / reference column)", [""] + headers, index=0)
        with lk2:
            return_header = st.selectbox("Return Column (value to return)", [""] + headers, index=0)
            optional_filter_header = st.selectbox("Optional Filter Column (optional)", [""] + headers, index=0)

        def apply_lookup_mapping(results: List[dict]) -> Tuple[int, int]:
            # re-run finalize + token replacement + revalidate, mirroring JS applyLookupMappingIfPresent
            warn = 0
            err = 0
            prof = analysis["profile"]
            for r in results:
                text = str(r["converted"] or "")
                has_template = "__LOOKUP_TEMPLATE__(" in text
                has_map_tokens = bool(re.search(r"__MAP_(RETURN|MATCH|LOOKUP)__", text, flags=re.I))
                if not (has_template or has_map_tokens):
                    continue

                ctx = Context(
                    col_header_map=analysis["col_header_map"],
                    target_row=r["row"],
                    header_row=analysis["header_row"],
                    profile_opts={**prof["opts"], "preferRowRefs": analysis["output_mode"] == "row"},
                    output_mode=analysis["output_mode"],
                    strict_mode=analysis["strict_mode"],
                    lookup_value_header=lookup_value_header,
                    match_header=match_header,
                    return_header=return_header,
                    optional_filter_header=optional_filter_header,
                )

                notes = Notes()
                f = text
                if has_template:
                    f = finalize_lookup_templates(f, ctx, notes)

                # Token substitution (if placeholders exist)
                lookup_val = f"[{lookup_value_header}]@row" if lookup_value_header else "__MAP_LOOKUP__"
                match_ref = f"{{{match_header}}}" if match_header else "__MAP_MATCH__"
                return_ref = f"{{{return_header}}}" if return_header else "__MAP_RETURN__"

                f = f.replace("__MAP_LOOKUP__", lookup_val).replace("__MAP_MATCH__", match_ref).replace("__MAP_RETURN__", return_ref)
                if optional_filter_header:
                    f = f.replace("__MAP_FILTER__", f"{{{optional_filter_header}}}")

                r["converted"] = f
                issues = validate_smartsheet_formula(f, ctx.profile_opts, ctx.strict_mode)
                r["errors"] = issues.errors
                r["warnings"] = issues.warnings + notes.warn
                r["info"] = issues.info + notes.info
                r["hasLookupTemplate"] = bool(re.search(r"__LOOKUP_TEMPLATE__\(|__MAP_(RETURN|MATCH|LOOKUP)__", f, flags=re.I))

            warn = sum(1 for rr in results if rr["warnings"])
            err = sum(1 for rr in results if rr["errors"])
            return warn, err

        apply_disabled = not analysis["has_lookup_or_placeholders"]
        if apply_disabled:
            st.info("No LOOKUP templates detected in this analysis.")
        else:
            btns = st.columns([1, 1, 2])
            with btns[0]:
                apply_btn = st.button("Apply Mapping to LOOKUP Templates", disabled=not (lookup_value_header and match_header and return_header))
            with btns[1]:
                clear_btn = st.button("Clear Mapping")
            if clear_btn:
                st.experimental_rerun()
            if apply_btn:
                w, e = apply_lookup_mapping(analysis["results"])
                analysis["warn_count"] = w
                analysis["err_count"] = e
                st.session_state.analysis = analysis
                st.success("Mapping applied. Re-export XLSX/CSV when ready.")

        st.divider()
        st.markdown("### Export")
        base = (analysis["workbook_name"] or "Workbook").replace(".xlsx", "")

        # Prepare exports on-demand
        colx, colc = st.columns(2)
        with colx:
            if st.button("Convert + Download XLSX"):
                out_bytes = build_smartsheet_ready_workbook(
                    in_wb=in_wb,
                    sheet_name=analysis["sheet_name"],
                    results=analysis["results"],
                    create_report=analysis["create_report"],
                    strict_mode=analysis["strict_mode"],
                )
                st.download_button(
                    "Download SmartsheetReady.xlsx",
                    data=out_bytes,
                    file_name=f"{base}__SmartsheetReady.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        with colc:
            if st.button("Export Paste-Ready CSV"):
                csv_bytes = build_paste_ready_csv(analysis["results"])
                st.download_button(
                    "Download SmartsheetFormulas_PasteReady.csv",
                    data=csv_bytes,
                    file_name=f"{base}__SmartsheetFormulas_PasteReady.csv",
                    mime="text/csv",
                )
